from datetime import datetime as dt
import pandas as pd
from openpyxl import Workbook

from .models import Checkpoint, Deal

#---------------------------------------------------------------------------------------------------------------------------------------------------------------
# IO Manager Class
#---------------------------------------------------------------------------------------------------------------------------------------------------------------
class IOManager():

	#------------------------------------------------------------------------------------------------------------------------------------------------------------
	# Deal Download
	#------------------------------------------------------------------------------------------------------------------------------------------------------------
	def deal_download(self, items, filepath):

		wb = Workbook()
		ws = wb.active
		ws.title = 'Deals'

		ws.cell(column=1, row=1).value = 'ID'
		ws.cell(column=2, row=1).value = 'RUC'
		ws.cell(column=3, row=1).value = 'CPN'
		ws.cell(column=4, row=1).value = 'Razon Social'
		ws.cell(column=5, row=1).value = 'Fecha Ganado'
		ws.cell(column=6, row=1).value = 'Comercial'
		ws.cell(column=7, row=1).value = 'Plan BSale'
		ws.cell(column=8, row=1).value = 'Categoria'
		ws.cell(column=9, row=1).value = 'Ejecutivo PEM'
		ws.cell(column=10, row=1).value = 'Etapa'
		ws.cell(column=11, row=1).value = 'Estado'
		ws.cell(column=12, row=1).value = 'Fecha Inicio PEM'
		ws.cell(column=13, row=1).value = 'Fecha Contacto Inicial'
		ws.cell(column=14, row=1).value = 'Fecha Pase a Produccion'
		ws.cell(column=15, row=1).value = 'Dias en PEM'
		ws.cell(column=16, row=1).value = 'Dias en Prod.'
		ws.cell(column=17, row=1).value = 'URL BSale'
		ws.cell(column=18, row=1).value = 'URL Cliente'
		ws.cell(column=19, row=1).value = 'Fecha de Baja'
		ws.cell(column=20, row=1).value = 'Razon de Baja'

		for row, item in enumerate(items):
			ws.cell(column=1, row=row+2).value = item.negocio_id
			ws.cell(column=2, row=row+2).value = item.ruc
			ws.cell(column=3, row=row+2).value = item.cpn
			ws.cell(column=4, row=row+2).value = item.razon_social
			ws.cell(column=5, row=row+2).value = item.fecha_ganado
			ws.cell(column=6, row=row+2).value = item.comercial
			ws.cell(column=7, row=row+2).value = item.plan_bsale
			ws.cell(column=8, row=row+2).value = item.categoria
			ws.cell(column=9, row=row+2).value = item.ejecutivo_pem
			ws.cell(column=10, row=row+2).value = item.etapa
			ws.cell(column=11, row=row+2).value = item.estado
			ws.cell(column=12, row=row+2).value = item.fecha_inicio_pem
			ws.cell(column=13, row=row+2).value = item.fecha_contacto_inicial
			ws.cell(column=14, row=row+2).value = item.fecha_pase_produccion
			ws.cell(column=15, row=row+2).value = item.dias_pem()
			ws.cell(column=16, row=row+2).value = item.dias_prod()
			ws.cell(column=17, row=row+2).value = item.url_bsale
			ws.cell(column=18, row=row+2).value = item.url_cliente
			ws.cell(column=19, row=row+2).value = item.fecha_baja
			ws.cell(column=20, row=row+2).value = item.razon_baja

		wb.save(filename = filepath)
		
	#------------------------------------------------------------------------------------------------------------------------------------------------------------
	# Deal Load
	#------------------------------------------------------------------------------------------------------------------------------------------------------------

	def deal_load(self, excel_file, db):

		df = pd.read_sql_table('deals', con=db.engine, columns=['negocio_id'])

		df_new = pd.read_excel(excel_file, 
							usecols = ["Negocio - ID",
										"Negocio - RUT/RUC/NIT",
										"Negocio - Título", 
										"Negocio - Propietario", 
										"Negocio - Servicio Contratado",
										"Negocio - Rubro/Actividad Económica",
										"Negocio - Fecha de ganado"],
								dtype = {"Negocio - RUT/RUC/NIT":str, "Negocio - Fecha de ganado":str})

		df_new.rename( columns = { "Negocio - ID":"negocio_id",
									"Negocio - RUT/RUC/NIT":"ruc",
									"Negocio - Título":"razon_social",
									"Negocio - Propietario":"comercial",
									"Negocio - Servicio Contratado":"plan_bsale",
									"Negocio - Rubro/Actividad Económica":"categoria",
									"Negocio - Fecha de ganado":"fecha_ganado"}, 
						inplace=True)
			

		df_new.fillna( value = {'ruc':'',
								'razon_social':'',
								'comercial':'',
								'plan_bsale':'',
								'categoria':'',
								'fecha_ganado':''}, 
						inplace=True)
			
		df_new.loc[:,'cpn'] = df_new.loc[:,'razon_social'].apply(lambda x: int(x[-6:-1]) if x[-6:-1].isdigit() else ( int(x[-5:]) if x[-5:].isdigit() else 0 ) )
		df_new.loc[:,'fecha_ganado'] = df_new.loc[:,'fecha_ganado'].apply(lambda x: x[0:10])
		df_new.loc[:,'comercial'] = df_new.loc[:,'comercial'].apply(lambda x: x.split(' ')[0].upper() + ' ' + x.split(' ')[1][0].upper())
			
		planes = {
					"Plan Estándar PV":"Estándar PV",
					"Módulo Ecommerce":"Módulo Ecommerce",
					"Plan Estándar TO":"Estándar TO",
					"Plan Básico":"Básico",
					"Plan Full":"Omnicanal"
		}

		df_new.loc[:,'plan_bsale'] = df_new.loc[:,'plan_bsale'].apply(lambda x: planes[x] if x in planes else '')

		# Merge

		df_merge = pd.merge(df, df_new, how='outer', indicator=True)

		df_merge = df_merge[ df_merge['_merge']=='right_only']

		df_merge.drop(columns=['_merge'], inplace=True)

		#df_merge.drop_duplicates(subset ='cpn', keep='first', inplace=True)

		if not df_merge.empty:
			df_merge.loc[:,'etapa'] = ''
			df_merge.loc[:,'estado'] = ''
			df_merge.loc[:,'ejecutivo_pem'] = ''
			df_merge.loc[:,'fecha_inicio_pem'] = ''
			df_merge.loc[:,'fecha_contacto_inicial'] = ''
			df_merge.loc[:,'fecha_pase_produccion'] = ''
			df_merge.loc[:,'url_bsale'] = ''
			df_merge.loc[:,'comentario'] = 'Nuevo'

			#--- Write to DB

			df_merge.to_sql('deals', con=db.engine, if_exists='append', index=False)

			#--- Add CheckPoints

			seguimientos = {
				'Seguimiento Pasa a Produccion':3,
				'Seguimiento dia 15':15,
				'Seguimiento dia 30':30,
				'Seguimiento dia 60':60,
				}

			for id in df_merge.loc[:,'negocio_id']:

				for nombre, dias in seguimientos.items():
					
					checkpoint = Checkpoint(nombre=nombre, 
											fecha='', 
											realizado=False, 
											comentario='', 
											fecha_realizado='', 
											estado='', 
											deal_id=id)

					db.session.add(checkpoint)

				db.session.commit()

	#------------------------------------------------------------------------------------------------------------------------------------------------------------
	# Deal Update
	#------------------------------------------------------------------------------------------------------------------------------------------------------------
	def get_file_cols(self, excel_file):

		df = pd.read_excel(excel_file)

		return( [ col for col in df.columns] )


	#------------------------------------------------------------------------------------------------------------------------------------------------------------
	# Deal Update
	#------------------------------------------------------------------------------------------------------------------------------------------------------------
	def deal_update(self, excel_file, db, import_columns):

		df = pd.read_excel(excel_file)

		df.set_index('ID', inplace=True)

		df.fillna('', inplace=True)

		deals_updated = []
		
		for id in df.index:

			deal = Deal.query.get(id)

			if deal:

				for col in import_columns:
					if df.loc[id, col] != '':
						if col == 'RUC':
							deal.set_value_by_col(col, str(df.loc[id, col]))
						elif col == 'CPN':
							deal.set_value_by_col(col, int(df.loc[id, col]))
						else:
							deal.set_value_by_col(col, df.loc[id, col])				

				deals_updated.append(deal)

		db.session.commit()

		return(deals_updated)

#---------------------------------------------------------------------------------------------------------------------------------------------------------------
# IO Manager Object
#---------------------------------------------------------------------------------------------------------------------------------------------------------------

io_manager = IOManager()

